import os
import argparse
from PyPDF2 import PdfReader
from docx import Document
import openpyxl

def extract_text_from_file(file_path):
    """Extract text based on the file extension."""
    _, file_extension = os.path.splitext(file_path)
    file_extension = file_extension.lower()

    extractors = {
        '.txt': extract_text_from_txt,
        '.pdf': extract_text_from_pdf,
        '.docx': extract_text_from_docx,
        '.xlsx': extract_text_from_xlsx,
    }

    if file_extension in extractors:
        return extractors[file_extension](file_path)
    elif file_extension in ['.py', '.java', '.js', '.html', '.css', '.json', '.xml', '.c', '.cpp']:
        return extract_text_from_code(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_extension}")

def extract_text_from_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def extract_text_from_pdf(file_path):
    text = ""
    with open(file_path, 'rb') as f:
        reader = PdfReader(f)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    return text

def extract_text_from_docx(file_path):
    text = ""
    doc = Document(file_path)
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text

def extract_text_from_xlsx(file_path):
    text = ""
    workbook = openpyxl.load_workbook(file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) if cell is not None else '' for cell in row])
            text += row_text + "\n"
    return text

def extract_text_from_code(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def main():
    parser = argparse.ArgumentParser(
        description="TFQ_tool: Extract text from various file formats including text, PDFs, Word documents, Excel sheets, and code files.",
        epilog="Examples:\n"
               "  python TFQ_tool.py sample.pdf\n"
               "  python TFQ_tool.py document.docx --output extracted.txt\n"
               "  python TFQ_tool.py file1.txt file2.txt --verbose",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("file_paths", nargs='*', help="Path(s) to the file(s) for text extraction")
    parser.add_argument("-o", "--output", type=str, help="File to save the extracted text (default: generated automatically)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Enable verbose output for debugging or detailed status")

    args = parser.parse_args()

    if not args.file_paths:
        parser.print_help()
        return

    for file_path in args.file_paths:
        try:
            if args.verbose:
                print(f"Processing file: {file_path}")
            extracted_text = extract_text_from_file(file_path)
            if args.output:
                output_file = args.output
            else:
                output_file = f"{os.path.splitext(os.path.basename(file_path))[0]}_extracted.txt"
            with open(output_file, 'w', encoding='utf-8') as out_file:
                out_file.write(extracted_text)
            print(f"Text extracted successfully. Saved to {output_file}.")
        except Exception as e:
            print(f"An error occurred while processing {file_path}: {e}")

if __name__ == "__main__":
    main()
